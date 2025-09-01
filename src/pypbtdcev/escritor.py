# ----------------------------
# -------- ESCRITOR ----------
# ----------------------------

import openpyxl
import pandas as pd


class EscritorPBTD01_v2:
    def __init__(self):
      # --- Mapa para Puertas ---
        self.mapa_escritura = {
            '3. Tablas Envolvente': {
                'puertas': {
                    # Las primeras 6 filas (12-17) no se tocan
                    'filas_por_defecto': 6,
                    'fila_inicio_modificable': 18,
                    'filas_editables_max': 6,  # Filas 18 a 23 inclusive
                    #   Solo incluimos las columnas que SÍ podemos editar
                    'columnas_modificables': {
                        'nombre': 'B',
                        'abreviatura': 'C',
                        'u_puerta_opaca_w_m2k': 'D',
                        'vidrio': 'E',
                        'porcentaje_vidrio': 'F',
                        'u_marco_w_m2k': 'G',
                        'porcentaje_marco': 'H'
                    }
                },
                # --- Mapa para Vidrios ---
                'vidrios': {
                    'fila_inicio': 32,  # Fila donde comienzan los datos modificables
                    'filas_editables_max': 11,  # Filas 32 a 42 inclusive
                    'columnas': {
                        'nombre': 'B',
                        'abreviatura': 'C',
                        'u_vidrio_w_m2k': 'D',
                        'fs_vidrio': 'E'
                    }
                },
                # --- MAPA PARA MARCOS VENTANA ---
                'marcos_ventana': {
                    'fila_inicio': 50,  # La escritura de datos modificables comienza en la fila 50
                    'filas_editables_max': 8,  # Filas 50 a 57 inclusive
                    'columnas': {
                        'nombre_tipo_marcos': 'B',
                        'abreviatura': 'C',
                        'ufr_w_m2k': 'D',
                        'fm': 'E'
                    }
                },
                # --- MAPA PARA Muros transmitancia ---
                'muros_transmitancia': {
                    'hoja': '3. Tablas Envolvente',
                    'fila_inicio': 61,
                    'filas_editables_max': 15,  # Filas 61 a 75 inclusive
                    'celdas_no_modificables': ['B61', 'C61', 'E61'],
                    'columnas': {
                        'nombre': 'B',
                        'abreviatura': 'C',
                        'tipologia_materialidad': 'D',
                        'u_w_m2k': 'E',
                        'espesor_muro_solido_cm': 'F',
                        'espesor_aislante_cm': 'G',
                        'posicion_aislacion': 'H'
                    }
                },
                # --- MAPA PARA Techos transmitancia ---
                'techos_transmitancia': {
                    'fila_inicio': 79,
                    'filas_editables_max': 4,  # Filas 79 a 82 inclusive
                    'celdas_no_modificables': ['B79', 'C79', 'D79'],
                    'columnas': {
                        'nombre': 'B', 'abreviatura': 'C', 'u_w_m2k': 'D',
                        'espesor_techo_solido_cm': 'F', 'espesor_aislante_cm': 'G',
                        'posicion_aislacion': 'H'
                    }
                },
                # --- MAPA PARA Pisos transmitancia ---
                'pisos_transmitancia': {
                    'fila_inicio': 87,
                    'filas_editables_max': 5,  # Filas 87 a 91 inclusive
                    'celdas_no_modificables': ['B87', 'C87', 'D87'],
                    'columnas': {
                        'nombre': 'B', 'abreviatura': 'C', 'u_piso_ventilado_w_m2k': 'D',
                        'aislacion_terreno_lambda_w_mk': 'E', 'aislacion_terreno_e_aislante_cm': 'F',
                        'refuerzo_vert_lambda_w_mk': 'G', 'refuerzo_vert_e_aislante_cm': 'H',
                        'refuerzo_vert_d_cm': 'I', 'refuerzo_horiz_lambda_w_mk': 'J',
                        'refuerzo_horiz_e_aislante_cm': 'K', 'refuerzo_horiz_d_cm': 'L',
                        'posicion_aislacion': 'M'
                    }
                }
            },
            'CEV-CEVE': {
                'datos_generales_proyecto': {
                    'celdas': {
                        'tipo_de_calificacion': 'E7',
                        'tipo_de_vivienda_calificacion': 'G7',
                        'region': 'E8',
                        'comuna': 'E9',
                        'zona_termica_proyecto': 'E10',
                        'dormitorios_de_la_vivienda': 'E11',
                        'identificacion_de_la_vivienda_a_evaluar': 'E13',
                        'nombre_del_proyecto': 'E14',
                        'direccion_de_la_vivienda': 'E15',
                        'tipo_de_vivienda': 'E16',
                        'rol_vivienda': 'E19',
                        'evaluador_energetico': 'E20',
                        'rol_registro_de_evaluadores': 'E21',
                        'rut_evaluador': 'E22',
                        # 'version_planilla':'E24', # celda no editable
                        'caso_interno_evaluador': 'E25',
                        'iteracion_evaluador': 'E26',
                        'solicitado_por': 'E28',
                        'rut_mandante': 'E29'
                    }
                },
                # --- Mapa para Elementos de la Envolvente ---
                'elementos_de_la_envolvente': {
                    'celdas': {
                        'muro_principal': 'E33', 'muro_secundario': 'E34', 'piso_principal': 'E35',
                        'techo_principal': 'E36', 'techo_secundario': 'E37', 'ventana_principal_vidrio': 'E38',
                        'ventana_principal_marco': 'O38', 'ventana_secundaria_vidrio': 'E39',
                        'ventana_secundaria_marco': 'O39', 'puerta_principal': 'E40'
                    }
                },
                # --- Mapa para Calefacción y ACS ---
                'calefaccion_y_acs': {
                    'celdas': {
                        'sistema_de_calefaccion': 'E44',
                        'sistema_de_agua_caliente': 'E45'
                    }
                },
                # --- Mapa para la tabla de Dimensiones ---
                'dimensiones_de_la_vivienda': {
                    'fila_inicio': 52,
                    'columnas': {
                        # OJO: Solo mapeamos las columnas que son datos de entrada.
                        # 'volumen_m3' y los totales son calculados por Excel.
                        'piso': 'C',
                        'area_m2': 'D',
                        'altura_m': 'E'
                    }
                },
                'area_y_coeficiente_muros': {
                    'fila_inicio': 66,
                    'filas_editables_max': 16,  # El lector extrae 16 filas
                    'columnas': {
                        'nombre_muro': 'D',
                        'angulo_azimut': 'E',
                        'area_m2': 'H',
                        'puente_termico_p01': 'K',
                        'puente_termico_p02': 'L',
                        'puente_termico_p03': 'M'
                    }
                },
                # --- Mapa para Puentes Térmicos Particulares ---
                'puentes_termicos_particulares': {
                    'fila_inicio': 87,
                    'filas_editables_max': 5,  # El lector extrae 5 filas
                    'columnas': {
                        'alojada_en_muro': 'D',
                        'azimut': 'E',
                        'elemento_perpendicular': 'G',
                        'aislacion': 'H',
                        'longitud_m': 'I'
                    }
                },
                # --- Mapa para Puertas ---
                'puertas': {
                    'fila_inicio': 96,
                    'filas_editables_max': 3,  # El lector extrae 3 filas
                    'columnas': {
                        # Mapeamos solo las columnas de entrada del usuario
                        'tipo_puerta': 'D',
                        'azimut': 'E',
                        'categoria_infiltracion': 'H',
                        'alto_m': 'K',
                        'ancho_m': 'L',
                        'fav1_d': 'P',
                        'fav1_l': 'Q',
                        'fav2_izquierda_p': 'R',
                        'fav2_izquierda_s': 'S',
                        'fav2_derecha_p': 'T',
                        'fav2_derecha_s': 'U',
                        'fav3_e': 'V',
                        'fav3_t': 'W',
                        'fav3_beta': 'X',
                        'fav3_alpha': 'Y'
                    }
                },
                # --- Mapa para Ventanas ---
                'ventanas': {
                    'fila_inicio': 103,
                    'filas_editables_max': 20,
                    'columnas': {
                        'tipo_ventana': 'D',
                        'azimut': 'E',
                        'elemento_envolvente': 'G',
                        'tipo_cierre': 'H',
                        'posicion_ventanal': 'I',
                        'aislacion_con_sin_retorno': 'J',
                        'alto_m': 'K',
                        'ancho_m': 'L',
                        'categoria_para_pt_y_infilt': 'M',
                        'tipo_marco': 'N',
                        'fav1_d': 'P',
                        'fav1_l': 'Q',
                        'fav2_izquierda_p': 'R',
                        'fav2_izquierda_s': 'S',
                        'fav2_derecha_p': 'T',
                        'fav2_derecha_s': 'U',
                        'fav3_e': 'V',
                        'fav3_t': 'W',
                        'fav3_beta': 'X',
                        'fav3_alpha': 'Y'

                    }
                },
                # --- Mapa para la sección de Obstrucciones ---
                'obstrucciones': {
                    # El mapa contiene las anclas (en formato de índice de pandas) para cada bloque.
                    # Ancla = Fila y Columna de la celda de la Orientación (ej: 'N' en D125)
                    'orientaciones': {
                        'N': (124, 4),  # Celda E125
                        'E': (124, 9),  # Celda J125
                        'S': (124, 14),  # Celda O125
                        'O': (124, 19),  # Celda T125
                        'NE': (136, 4),  # Celda E137
                        'SE': (136, 9),  # Celda J137
                        'SO': (136, 14),  # Celda O137
                        'NO': (136, 19)  # Celda T137
                    },
                    'columnas_tabla': ['division', 'a_m', 'b_m', 'd_m']
                },
                # --- Mapa para la tabla de Techos ---
                'techos': {
                    'fila_inicio': 150,
                    'filas_editables_max': 5,  # El lector extrae 5 filas
                    'columnas': {
                        # Mapeamos las columnas de entrada del usuario
                        'techos': 'D',
                        'densidad_techo': 'F',
                        'area_m2': 'G',
                        'camaras_de_aire': 'K',
                        'tipo_de_cubierta': 'L'
                    }
                },
                # --- Mapa para la tabla de Pisos ---
                'pisos': {
                    'fila_inicio': 158,
                    'filas_editables_max': 4,  # El lector extrae 4 filas
                    'columnas': {
                        'piso': 'D',
                        'densidad_piso': 'F',
                        'area_m2': 'G',
                        'perimetro_contacto_terreno_m': 'K',
                        'piso_ventilado': 'L'

                    }
                },
                # --- Mapa para la sección de Condiciones de Uso ---
                'condiciones_de_uso': {
                    'infiltraciones': {
                        'celdas': {
                            'cuenta_con_ensayo_presurizacion': 'F217',
                            'valor_ensayo_presurizacion_rah_a_50pa': 'F219',
                            'cantidad_ductos_ventilacion': 'F221',
                            'cantidad_celosias': 'F223'
                        }
                    },
                    'ventilacion': {
                        'celdas': {
                            'ventilacion_mecanica_vm': 'E227',
                            'eficiencia_recuperador_calor_porc': 'F229',
                            'tiene_sensor_co2': 'F232',
                            'rah_segun_memoria_calculo': 'F234'
                        }
                    }
                }


            }
        }

    def _escribir_tabla_puertas(self, ws, datos_puertas):
        """
        Escribe los datos de la tabla Puertas, respetando las filas y columnas no editables.
        """
        print("Escribiendo datos de la tabla 'Puertas'...")
        mapa = self.mapa_escritura['3. Tablas Envolvente']['puertas']
        fila_actual = mapa['fila_inicio_modificable']
        max_rows = mapa['filas_editables_max']

        # Omitimos las filas por defecto de nuestra lista de datos
        datos_modificables = datos_puertas[mapa['filas_por_defecto']:]

        # Filtramos para escribir solo las filas que tienen una abreviatura
        datos_a_escribir = [fila for fila in datos_modificables if fila.get(
            'abreviatura') is not None]

        # Verificación limite de filas editables
        if len(datos_a_escribir) > max_rows:
            print(
                f"    ⚠️ ADVERTENCIA: Se proporcionaron {len(datos_a_escribir)} registros para 'Puertas', pero solo hay espacio para {max_rows}.")
            print(
                f"    -> Se escribirán solo los primeros {max_rows} registros.")
            # Truncamos la lista para que solo contenga los datos que caben
            datos_a_escribir = datos_a_escribir[:max_rows]

        for registro_puerta in datos_a_escribir:
            # Iteramos SOLAMENTE sobre las columnas que definimos como modificables
            for key, col in mapa['columnas_modificables'].items():
                celda = f"{col}{fila_actual}"
                valor = registro_puerta.get(key)
                if valor is not None:
                    ws[celda] = valor
            fila_actual += 1
        print(
            f" -> {len(datos_a_escribir)} registros de 'Puertas' escritos en filas modificables.")

    def _escribir_tabla_vidrios(self, ws, datos_vidrios):
        print("Escribiendo datos de la tabla 'Vidrios'...")
        mapa = self.mapa_escritura['3. Tablas Envolvente']['vidrios']
        fila_actual = mapa['fila_inicio']
        max_rows = mapa['filas_editables_max']

        # 1. Omitimos las primeras 5 filas del diccionario (valores por defecto)
        datos_modificables = datos_vidrios[5:]

        # 2. Filtramos la lista para quedarnos solo con las filas que tienen datos
        # Asumimos que si no tiene 'abreviatura', es una fila vacía que no debe escribirse.
        datos_a_escribir = [fila for fila in datos_modificables if fila.get(
            'abreviatura') is not None]

        # Verificación limite de filas editables
        if len(datos_a_escribir) > max_rows:
            print(
                f"    ⚠️ ADVERTENCIA: Se proporcionaron {len(datos_a_escribir)} registros para 'Vidrios', pero solo hay espacio para {max_rows}.")
            print(
                f"    -> Se escribirán solo los primeros {max_rows} registros.")
            # Truncamos la lista para que solo contenga los datos que caben
            datos_a_escribir = datos_a_escribir[:max_rows]

        for registro_vidrio in datos_a_escribir:
            for key, col in mapa['columnas'].items():
                celda = f"{col}{fila_actual}"
                valor = registro_vidrio.get(key)
                if valor is not None:
                    ws[celda] = valor
            fila_actual += 1
        print(f" -> {len(datos_a_escribir)} registros de 'Vidrios' escritos.")

    def _escribir_tabla_marcos_ventana(self, ws, datos_marcos):
        """
        Escribe los datos de la tabla Marcos Ventana en la hoja de cálculo.
        """
        print("Escribiendo datos de la tabla 'Marcos Ventana'...")
        mapa = self.mapa_escritura['3. Tablas Envolvente']['marcos_ventana']
        fila_actual = mapa['fila_inicio']
        max_rows = mapa['filas_editables_max']

        # 1. Omitimos las primeras 4 filas (valores por defecto)
        datos_modificables = datos_marcos[4:]

        # 2. Filtramos para escribir solo las filas con datos
        # Asumimos que si no tiene 'abreviatura', es una fila vacía que no debe escribirse.
        datos_a_escribir = [fila for fila in datos_modificables if fila.get(
            'abreviatura') is not None]

        # Verificación limite de filas editables
        if len(datos_a_escribir) > max_rows:
            print(
                f"    ⚠️ ADVERTENCIA: Se proporcionaron {len(datos_a_escribir)} registros para 'Marcos Ventanas', pero solo hay espacio para {max_rows}.")
            print(
                f"    -> Se escribirán solo los primeros {max_rows} registros.")
            # Truncamos la lista para que solo contenga los datos que caben
            datos_a_escribir = datos_a_escribir[:max_rows]

        for registro_marco in datos_a_escribir:
            for key, col in mapa['columnas'].items():
                celda = f"{col}{fila_actual}"
                valor = registro_marco.get(key)
                if valor is not None:
                    ws[celda] = valor
            fila_actual += 1
        print(
            f" -> {len(datos_a_escribir)} registros de 'Marcos Ventana' escritos.")

    def _escribir_tabla_muros(self, ws, datos_muros):
        """
        Escribe los datos de la tabla Muros transmitancia,
        omitiendo las celdas no modificables.
        """
        print("Escribiendo datos de la tabla 'Muros transmitancia'...")
        mapa = self.mapa_escritura['3. Tablas Envolvente']['muros_transmitancia']
        fila_actual = mapa['fila_inicio']
        celdas_no_modificables = mapa['celdas_no_modificables']
        max_rows = mapa['filas_editables_max']

        # Filtramos para escribir solo las filas que tienen una abreviatura
        datos_a_escribir = [fila for fila in datos_muros if fila.get(
            'abreviatura') is not None]

        # Verificación limite de filas editables
        if len(datos_a_escribir) > max_rows:
            print(
                f"    ⚠️ ADVERTENCIA: Se proporcionaron {len(datos_a_escribir)} registros para 'Muros transmitancia', pero solo hay espacio para {max_rows}.")
            print(
                f"    -> Se escribirán solo los primeros {max_rows} registros.")
            # Truncamos la lista para que solo contenga los datos que caben
            datos_a_escribir = datos_a_escribir[:max_rows]

        for i, registro_muro in enumerate(datos_a_escribir):
            fila_actual = mapa['fila_inicio'] + i
            for key, col in mapa['columnas'].items():
                celda = f"{col}{fila_actual}"

                # --- LÓGICA DE EXCEPCIÓN ---
                # Si la celda está en la lista de no modificables, la saltamos.
                if celda in celdas_no_modificables:
                    continue

                valor = registro_muro.get(key)
                if valor is not None:
                    ws[celda] = valor
        print(
            f" -> {len(datos_a_escribir)} registros de 'Muros transmitancia' escritos.")

    def _escribir_tabla_techos(self, ws, datos_techos):
        """
        Escribe los datos de la tabla Techos transmitancia,
        omitiendo las celdas no modificables.
        """
        print("Escribiendo datos de la tabla 'Techos transmitancia'...")
        mapa = self.mapa_escritura['3. Tablas Envolvente']['techos_transmitancia']
        celdas_no_modificables = mapa['celdas_no_modificables']
        max_rows = mapa['filas_editables_max']

        # Filtramos para escribir solo las filas que tienen una abreviatura
        datos_a_escribir = [fila for fila in datos_techos if fila.get(
            'abreviatura') is not None]

        # Verificación limite de filas editables
        if len(datos_a_escribir) > max_rows:
            print(
                f"    ⚠️ ADVERTENCIA: Se proporcionaron {len(datos_a_escribir)} registros para 'Techos transmitancia', pero solo hay espacio para {max_rows}.")
            print(
                f"    -> Se escribirán solo los primeros {max_rows} registros.")
            # Truncamos la lista para que solo contenga los datos que caben
            datos_a_escribir = datos_a_escribir[:max_rows]

        for i, registro_techo in enumerate(datos_a_escribir):
            fila_actual = mapa['fila_inicio'] + i
            for key, col in mapa['columnas'].items():
                celda = f"{col}{fila_actual}"

                # Lógica para saltarse las celdas no modificables
                if celda in celdas_no_modificables:
                    continue

                valor = registro_techo.get(key)
                if valor is not None:
                    ws[celda] = valor
        print(
            f" -> {len(datos_a_escribir)} registros de 'Techos transmitancia' escritos.")

    def _escribir_tabla_pisos(self, ws, datos_pisos):
        """
        Escribe los datos de la tabla Pisos transmitancia,
        omitiendo las celdas no modificables.
        """
        print("Escribiendo datos de la tabla 'Pisos transmitancia'...")
        mapa = self.mapa_escritura['3. Tablas Envolvente']['pisos_transmitancia']
        celdas_no_modificables = mapa['celdas_no_modificables']
        max_rows = mapa['filas_editables_max']

        datos_a_escribir = [fila for fila in datos_pisos if fila.get(
            'abreviatura') is not None]

        # Verificación limite de filas editables
        if len(datos_a_escribir) > max_rows:
            print(
                f"    ⚠️ ADVERTENCIA: Se proporcionaron {len(datos_a_escribir)} registros para 'Pisos transmitancia', pero solo hay espacio para {max_rows}.")
            print(
                f"    -> Se escribirán solo los primeros {max_rows} registros.")
            # Truncamos la lista para que solo contenga los datos que caben
            datos_a_escribir = datos_a_escribir[:max_rows]

        for i, registro_piso in enumerate(datos_a_escribir):
            fila_actual = mapa['fila_inicio'] + i
            for key, col in mapa['columnas'].items():
                celda = f"{col}{fila_actual}"

                if celda in celdas_no_modificables:
                    continue

                valor = registro_piso.get(key)
                if valor is not None:
                    ws[celda] = valor
        print(
            f" -> {len(datos_a_escribir)} registros de 'Pisos transmitancia' escritos.")

    def _escribir_datos_clave_valor(self, ws, datos_seccion, mapa_seccion):
        """Función genérica para escribir datos de secciones tipo clave-valor."""
        for clave, celda in mapa_seccion['celdas'].items():
            valor = datos_seccion.get(clave)
            if pd.notna(valor):
                ws[celda] = valor

    def _escribir_tabla_dimensiones_cev(self, ws, datos_dimensiones, mapa_seccion):
        print("Escribiendo datos de la tabla 'Dimensiones de la Vivienda'...")
        fila_actual = mapa_seccion['fila_inicio']
        lista_pisos = datos_dimensiones.get('pisos', [])

        for registro in lista_pisos:
            if registro.get('piso') is not None:
                for key, col in mapa_seccion['columnas'].items():
                    valor = registro.get(key)
                    if pd.notna(valor):
                        ws[f"{col}{fila_actual}"] = valor
                fila_actual += 1
        print(f" -> Datos de 'Dimensiones' escritos.")

    def _escribir_tabla_cev(self, ws, nombre_seccion, datos_tabla, mapa_seccion):
        """
        Función genérica para escribir las tablas de la hoja CEV-CEVE.
        """
        print(f"Escribiendo tabla '{nombre_seccion}'...")
        fila_actual = mapa_seccion['fila_inicio']
        max_rows = mapa_seccion.get('filas_editables_max', len(datos_tabla))

        # Filtramos para escribir solo las filas que contienen datos
        # Asumimos que si no tiene la primera columna del mapa, es una fila vacía
        id_columna_check = list(mapa_seccion['columnas'].keys())[0]
        datos_a_escribir = [fila for fila in datos_tabla if fila.get(
            id_columna_check) is not None]

        if len(datos_a_escribir) > max_rows:
            print(
                f"    ⚠️ ADVERTENCIA: Se proporcionaron {len(datos_a_escribir)} registros para '{nombre_seccion}', pero solo hay espacio para {max_rows}.")
            datos_a_escribir = datos_a_escribir[:max_rows]

        for registro in datos_a_escribir:
            for key, col in mapa_seccion['columnas'].items():
                valor = registro.get(key)
                if pd.notna(valor):
                    ws[f"{col}{fila_actual}"] = valor
            fila_actual += 1
        print(
            f" -> {len(datos_a_escribir)} registros de '{nombre_seccion}' escritos.")

    def _escribir_seccion_obstrucciones(self, ws, datos_obstrucciones, mapa_seccion):
        print("Escribiendo sección 'Obstrucciones'...")
        mapa_orientaciones = mapa_seccion['orientaciones']

        # Iteramos sobre cada orientación (N, E, S, O, etc.)
        for orientacion, anclas in mapa_orientaciones.items():
            if orientacion in datos_obstrucciones:
                datos_bloque = datos_obstrucciones[orientacion]
                ancla_fila, ancla_col = anclas

                # Escribir valores individuales (soloamente azimut)
                azimut = datos_bloque.get('azimut_rango')
                if azimut is not None:
                    ws.cell(row=ancla_fila + 1,
                            column=ancla_col + 3).value = azimut

                # Escribir la tabla de 8 obstrucciones
                fila_inicio_tabla = ancla_fila + 3
                for i, detalle in enumerate(datos_bloque.get('obstrucciones_detalle', [])):
                    fila_actual = fila_inicio_tabla + i
                    # Escribimos los 4 valores de la tabla
                    ws.cell(row=fila_actual, column=ancla_col +
                            1).value = detalle.get('division')
                    ws.cell(row=fila_actual, column=ancla_col +
                            2).value = detalle.get('a_m')
                    ws.cell(row=fila_actual, column=ancla_col +
                            3).value = detalle.get('b_m')
                    ws.cell(row=fila_actual, column=ancla_col +
                            4).value = detalle.get('d_m')
        print(" -> Datos de 'Obstrucciones' escritos.")

    def _escribir_seccion_condiciones_uso(self, ws, datos_seccion, mapa_seccion):
        """
        Orquesta la escritura de la sección compleja 'condiciones_de_uso'.
        """
        print("Escribiendo sección 'Condiciones de Uso'...")

        # Escribir sub-sección de infiltraciones
        if 'infiltraciones' in datos_seccion and 'infiltraciones' in mapa_seccion:
            print(" -> Escribiendo sub-sección 'infiltraciones'...")
            self._escribir_datos_clave_valor(
                ws,
                datos_seccion['infiltraciones'],
                mapa_seccion['infiltraciones']
            )

        # Escribir sub-sección de ventilación
        if 'ventilacion' in datos_seccion and 'ventilacion' in mapa_seccion:
            print(" -> Escribiendo sub-sección 'ventilacion'...")
            self._escribir_datos_clave_valor(
                ws,
                datos_seccion['ventilacion'],
                mapa_seccion['ventilacion']
            )

    def crear_nueva_planilla(self, ruta_plantilla, ruta_salida, datos):
        """
        Crea una nueva planilla a partir de una plantilla y escribe los datos modificados.
        """
        try:
            # Cargar el workbook existente, manteniendo las macros
            print(f"Cargando plantilla desde '{ruta_plantilla}'...")
            wb = openpyxl.load_workbook(ruta_plantilla, keep_vba=True)
            print(" -> Plantilla cargada.")

            # --- Escribir en la hoja '3. Tablas Envolvente' ---
            sheet_name_envolvente = '3. Tablas Envolvente'
            if sheet_name_envolvente in datos and sheet_name_envolvente in wb.sheetnames:
                print(
                    f"\n--- Iniciando escritura en hoja '{sheet_name_envolvente}' ---")
                hoja_envolvente = wb[sheet_name_envolvente]
                datos_envolvente = datos[sheet_name_envolvente]

                # Llama a la función específica para escribir la tabla de puertas
                if 'puertas' in datos_envolvente:
                    self._escribir_tabla_puertas(
                        hoja_envolvente, datos_envolvente['puertas'])

                # Llama a la función específica para escribir la tabla de vidrios
                if 'vidrios' in datos_envolvente:
                    self._escribir_tabla_vidrios(
                        hoja_envolvente, datos_envolvente['vidrios'])

                # Llama a la función para escribir la tabla de marcos ventana
                if 'marcos_ventana' in datos_envolvente:
                    self._escribir_tabla_marcos_ventana(
                        hoja_envolvente, datos_envolvente['marcos_ventana'])

                # Llama a la función para escribir la tabla de muros transmitancia
                if 'muros_transmitancia' in datos_envolvente:
                    self._escribir_tabla_muros(
                        hoja_envolvente, datos_envolvente['muros_transmitancia'])

                # Llama a la función para escribir la tabla de techos transmitancia
                if 'techos_transmitancia' in datos_envolvente:
                    self._escribir_tabla_techos(
                        hoja_envolvente, datos_envolvente['techos_transmitancia'])

                # Llama a la función para escribir la tabla de pisos transmitancia
                if 'pisos_transmitancia' in datos_envolvente:
                    self._escribir_tabla_pisos(
                        hoja_envolvente, datos_envolvente['pisos_transmitancia'])

            # --- Escribir en la hoja 'CEV-CEVE' ---
            sheet_name_cev = 'CEV-CEVE'
            if sheet_name_cev in datos and sheet_name_cev in wb.sheetnames:
                print(
                    f"\n--- Iniciando escritura en hoja '{sheet_name_cev}' ---")
                hoja_cev = wb[sheet_name_cev]
                datos_cev = datos[sheet_name_cev]
                mapa_cev = self.mapa_escritura[sheet_name_cev]

                # Escribir la sección 'datos_generales_proyecto'
                seccion_gral = 'datos_generales_proyecto'
                if seccion_gral in datos_cev and seccion_gral in mapa_cev:
                    print(f"-> Escribiendo sección '{seccion_gral}'...")
                    self._escribir_datos_clave_valor(
                        hoja_cev, datos_cev[seccion_gral], mapa_cev[seccion_gral])
                else:
                    print(
                        f"-> OMITIDO: No se encontró la sección '{seccion_gral}' en los datos o en el mapa.")

                # Escribir la sección 'elementos_de_la_envolvente'
                seccion_env = 'elementos_de_la_envolvente'
                if seccion_env in datos_cev and seccion_env in mapa_cev:
                    print(f"-> Escribiendo sección '{seccion_env}'...")
                    self._escribir_datos_clave_valor(
                        hoja_cev, datos_cev[seccion_env], mapa_cev[seccion_env])
                else:
                    print(
                        f"-> OMITIDO: No se encontró la sección '{seccion_env}' en los datos o en el mapa.")

                # Escribir la sección 'elementos_de_la_envolvente'
                seccion_cal = 'calefaccion_y_acs'
                if seccion_cal in datos_cev and seccion_cal in mapa_cev:
                    self._escribir_datos_clave_valor(
                        hoja_cev, datos_cev[seccion_cal], mapa_cev[seccion_cal])
                else:
                    print(
                        f"-> OMITIDO: No se encontró la sección '{seccion_cal}' en los datos o en el mapa.")

                # Escribir la sección 'dimensiones_de_la_vivienda'
                seccion_dim = 'dimensiones_de_la_vivienda'
                if seccion_dim in datos_cev and seccion_dim in mapa_cev:
                    self._escribir_tabla_dimensiones_cev(
                        hoja_cev, datos_cev[seccion_dim], mapa_cev[seccion_dim])
                else:
                    print(
                        f"-> OMITIDO: No se encontró la sección '{seccion_dim}' en los datos o en el mapa.")

                # Escribir la sección 'area_y_coeficiente_muros'
                seccion_muros = 'area_y_coeficiente_muros'
                if seccion_muros in datos_cev and seccion_muros in mapa_cev:
                    self._escribir_tabla_cev(
                        hoja_cev, seccion_muros, datos_cev[seccion_muros], mapa_cev[seccion_muros])
                else:
                    print(
                        f"-> OMITIDO: No se encontró la sección '{seccion_muros}' en los datos o en el mapa.")

                # Escribir la sección 'puentes_termicos_particulares'
                seccion_pt = 'puentes_termicos_particulares'
                if seccion_pt in datos_cev and seccion_pt in mapa_cev:
                    self._escribir_tabla_cev(
                        hoja_cev, seccion_pt, datos_cev[seccion_pt], mapa_cev[seccion_pt])
                else:
                    print(
                        f"-> OMITIDO: No se encontró la sección '{seccion_pt}' en los datos o en el mapa.")

                # Escribir la sección 'puertas'
                seccion_puertas = 'puertas'
                if seccion_puertas in datos_cev and seccion_puertas in mapa_cev:
                    self._escribir_tabla_cev(
                        hoja_cev, seccion_puertas, datos_cev[seccion_puertas], mapa_cev[seccion_puertas])
                else:
                    print(
                        f"-> OMITIDO: No se encontró la sección '{seccion_puertas}' en los datos o en el mapa.")

                # Escribir la sección 'ventanas'
                seccion_ventanas = 'ventanas'
                if seccion_ventanas in datos_cev and seccion_ventanas in mapa_cev:
                    self._escribir_tabla_cev(
                        hoja_cev, seccion_ventanas, datos_cev[seccion_ventanas], mapa_cev[seccion_ventanas])
                else:
                    print(
                        f"-> OMITIDO: No se encontró la sección '{seccion_ventanas}' en los datos o en el mapa.")

                # Escribir la sección 'obstrucciones'
                seccion_obs = 'obstrucciones'
                if seccion_obs in datos_cev and seccion_obs in mapa_cev:
                    self._escribir_seccion_obstrucciones(
                        hoja_cev, datos_cev[seccion_obs], mapa_cev[seccion_obs])
                else:
                    print(
                        f"-> OMITIDO: No se encontró la sección '{seccion_obs}' en los datos o en el mapa.")

                # Escribir la sección 'techos'
                seccion_techos = 'techos'
                if seccion_techos in datos_cev and seccion_techos in mapa_cev:
                    self._escribir_tabla_cev(
                        hoja_cev, seccion_techos, datos_cev[seccion_techos], mapa_cev[seccion_techos])
                else:
                    print(
                        f"-> OMITIDO: No se encontró la sección '{seccion_techos}' en los datos o en el mapa.")

                 # Escribir la sección 'pisos'
                seccion_pisos = 'pisos'
                if seccion_pisos in datos_cev and seccion_pisos in mapa_cev:
                    # Usamos 'piso' como la columna para verificar si la fila está vacía
                    self._escribir_tabla_cev(
                        hoja_cev, seccion_pisos, datos_cev[seccion_pisos], mapa_cev[seccion_pisos])
                else:
                    print(
                        f"-> OMITIDO: No se encontró la sección '{seccion_pisos}' en los datos o en el mapa.")

                # Escribir la sección 'condiciones_de_uso': 'infiltraciones' y 'ventilacion'
                seccion_uso = 'condiciones_de_uso'
                if seccion_uso in datos_cev and seccion_uso in mapa_cev:
                    self._escribir_seccion_condiciones_uso(
                        hoja_cev, datos_cev[seccion_uso], mapa_cev[seccion_uso])
                else:
                    print(
                        f"-> OMITIDO: No se encontró la sección '{seccion_uso}' en los datos o en el mapa.")

            else:
                print(
                    f"\n OMITIDO: No se encontró la hoja '{sheet_name_cev}' en los datos de entrada o en la planilla.")

            # Guardar el nuevo archivo
            wb.save(ruta_salida)
            print(f"✅ ¡Éxito! Planilla guardada en '{ruta_salida}'")

        except Exception as e:
            print(f"❌ Ocurrió un error al escribir el archivo: {e}")
